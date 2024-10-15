import os, uuid
# import xlsxwriter
from openpyxl import Workbook
import io
from datetime import datetime, timedelta
from order.models import *
from client.models import *
from trainer.models import *

class Excel:
    def __init__(self):
        pass

    def save_list(self):
        """   
        Parameters
        ----------
        kwargs
            불러와진 페이지에서 전달 받은 데이터
    
        Note
        ----
        주문 리스트를 엑셀 파일로 다운받는 함수
        """
        # idx_title = ['결제날짜', '배송시작날짜', '매장명', '담당 트레이너명', '회원명', '배송지주소', '상세주소', '공동현관 비밀번호', '연락처', '결제금액', 
        #              '추가 블록시 금액', '합계금액', '배송,픽업', '주차', '식사(1식,2식)', '요일', '날짜', '식단', '베이스', '단백질', '채소', '소스', 
        #              '추가단백질(1)', '추가단백질(2)', '추가야채(1)', '추가야채(2)', '추가소스']
        # one_order_list = Order.objects.filter(day_cnt=1).order_by('delivery_dt')
        # two_order_list = Order.objects.filter(day_cnt=2).order_by('delivery_dt')

        # output = BytesIO()
        # workbook = xlsxwriter.Workbook(output)
        # # workbook = xlsxwriter.Workbook("/Users/choco/다운로드")
        # # workbook = xlsxwriter.Workbook(response)
        # worksheet = workbook.add_worksheet('주문 목록')
        # bold = workbook.add_format({'bold': True})
        # center = workbook.add_format({'align': 'center', 'valign': 'vcenter'})

        # row = 0
        # # worksheet.write(row, 0, '1일 1식 회원', bold)
        # for idx, val in enumerate(idx_title):
        #     # idx = idx + 1
        #     worksheet.write(row, idx, val, bold)
        # row += 1

        # for order_cnt, order_info in enumerate(one_order_list):
        #     for week_cnt, week_info in enumerate(Order_Week.objects.filter(order=order_info).order_by('week')):
        #         for day_cnt, day_info in enumerate(Order_Detail.objects.filter(order_week=week_info).order_by('day')):
        #             worksheet.write(row, 0, Payment.objects.filter(order=order_info).latest('request_dt').request_dt.strftime("%Y.%m.%d"))
        #             print(Payment.objects.filter(order=order_info).latest('request_dt').request_dt.strftime("%Y.%m.%d"))
        #             worksheet.write(row, 1, order_info.delivery_dt.strftime("%Y.%m.%d"))
        #             print(order_info.delivery_dt.strftime("%Y.%m.%d"))
        #             worksheet.write(row, 2, order_info.client.trainer.gym.name)
        #             print(order_info.client.trainer.gym.name)
        #             worksheet.write(row, 3, order_info.client.trainer.name)
        #             print(order_info.client.trainer.name)
        #             worksheet.write(row, 4, order_info.client.name)
        #             print(order_info.client.name)
        #             worksheet.write(row, 5, Delivery.objects.get(client_id=order_info.client.client_id).address)
        #             print(Delivery.objects.get(client_id=order_info.client.client_id).address)
        #             worksheet.write(row, 6, Delivery.objects.get(client_id=order_info.client.client_id).address_detail)
        #             print(Delivery.objects.get(client_id=order_info.client.client_id).address_detail)
        #             if Delivery.objects.get(client_id=order_info.client.client_id).doorlock == 0:
        #                 doorlock = ''
        #             else:
        #                 doorlock = Delivery.objects.get(client_id=order_info.client.client_id).doorlock
        #             worksheet.write(row, 7, doorlock)
        #             print(Delivery.objects.get(client_id=order_info.client.client_id).doorlock)
        #             if order_info.client.contact[0] == '1':
        #                 default = str(order_info.client.contact)
        #                 contact = '0' + default[:2] + '-' + default[2:6] + '-' + default[6:]
        #             else:
        #                 contact = order_info.client.contact
        #             worksheet.write(row, 8, contact)
        #             print(contact)
        #             worksheet.write(row, 9, order_info.default_amount)
        #             print(order_info.default_amount)
        #             worksheet.write(row, 10, order_info.added_amount)
        #             print(order_info.added_amount)
        #             worksheet.write(row, 11, order_info.total_amount)
        #             print(order_info.total_amount)
        #             if order_info.is_delivery == True:
        #                 pickup = '픽업'
        #             else:
        #                 pickup = '배송'
        #             worksheet.write(row, 12, pickup)
        #             print(pickup)
        #             worksheet.write(row, 13, week_cnt)
        #             print(week_cnt)
        #             worksheet.write(row, 14, order_info.day_cnt)
        #             print(order_info.day_cnt)
        #             worksheet.write(row, 15, Day(day_info.day).label)
        #             print(Day(day_info.day).label)
        #             date = order_info.delivery_dt + timedelta(days=day_cnt)
        #             worksheet.write(row, 16, date.strftime("%Y.%m.%d"))
        #             print(date.strftime("%Y.%m.%d"))
        #             worksheet.write(row, 17, week_info.meal_group)
        #             print(week_info.meal_group)
        #             worksheet.write(row, 18, day_info.base_util.code)
        #             print(day_info.base_util.code)
        #             worksheet.write(row, 19, day_info.pro_util.code)
        #             print(day_info.pro_util.code)
        #             worksheet.write(row, 20, day_info.veg_util.code)
        #             print(day_info.veg_util.code)
        #             worksheet.write(row, 21, day_info.flavor_util.code)
        #             print(day_info.flavor_util.code)

        #             try:
        #                 for idx, add_pro_info in enumerate(Add_Pro.objects.filter(order_detail=day_info)):
        #                     cnt = 23 + idx
        #                     worksheet.write(row, cnt, Add_Pro.objects.get(add_pro_id=add_pro_info.add_pro_id).pro_util.code)
        #                     print(Add_Pro.objects.get(add_pro_id=add_pro_info.add_pro_id).pro_util.code)
        #             except:
        #                 pass
                    
        #             try:
        #                 for idx, add_veg_info in enumerate(Add_Veg.objects.filter(order_detail=day_info)):
        #                     cnt = 25 + idx
        #                     worksheet.write(row, cnt, Add_Veg.objects.get(add_veg_id=add_veg_info.add_veg_id).veg_util.code)
        #                     print(Add_Veg.objects.get(add_veg_id=add_veg_info.add_veg_id).veg_util.code)
        #             except:
        #                 pass
                    
        #             try:
        #                 worksheet.write(row, 27, Add_Flavor.objects.get(order_detail=day_info).flavor_util.code)
        #                 print(Add_Flavor.objects.get(order_detail=day_info).flavor_util.code)
        #             except:
        #                 pass
        #             row += 1
        # workbook.close()
        # output.seek(0)

        # return output

        # output = BytesIO()
        idx_title = ['결제날짜', '배송시작날짜', '매장명', '담당 트레이너명', '회원명', '배송지주소', '상세주소', '공동현관 비밀번호', '연락처', '결제금액', 
                     '추가 블록시 금액', '합계금액', '배송,픽업', '주차', '식사(1식,2식)', '요일', '날짜', '식단', '베이스', '단백질', '채소', '소스', 
                     '추가단백질(1)', '추가단백질(2)', '추가야채(1)', '추가야채(2)', '추가소스']
        one_order_list = Order.objects.filter(client_id__in=Client.objects.filter(trainer_id__in=Trainer.objects.filter(gym_id=Gym.objects.get(name='COZY원흥').gym_id).values_list('trainer_id', flat=True)).exclude(name='테스트').values_list('client_id', flat=True)).order_by('delivery_dt')
        two_order_list = Order.objects.filter(day_cnt=2).order_by('delivery_dt')

        workbook = Workbook()

        worksheet = workbook.create_sheet('주문 목록')
        worksheet = workbook.active

        # bold = workbook.add_format({'bold': True})
        # center = workbook.add_format({'align': 'center', 'valign': 'vcenter'})

        worksheet.append(idx_title)

        for order_cnt, order_info in enumerate(one_order_list):
            print(order_info.order_id, order_info.create_dt)
            try:
                if Payment.objects.filter(order=order_info).exists():
                    pass
                else:
                    Order_Detail.objects.filter(order_week_id__in=Order_Week.objects.filter(order=order_info).values_list('order_week_id', flat=True)).delete()
                    Order_Week.objects.filter(order=order_info).delete()
                    Order.objects.filter(order_id=order_info.order_id).delete()
                    continue
            except:
                Order_Detail.objects.filter(order_week_id__in=Order_Week.objects.filter(order=order_info).values_list('order_week_id', flat=True)).delete()
                Order_Week.objects.filter(order=order_info).delete()
                Order.objects.filter(order_id=order_info.order_id).delete()
                continue
            for week_cnt, week_info in enumerate(Order_Week.objects.filter(order=order_info).order_by('week')):
                for day_cnt, day_info in enumerate(Order_Detail.objects.filter(order_week=week_info).order_by('day')):
                    data = list()
                    print('making data list')
                    data.append(Payment.objects.filter(order=order_info).latest('request_dt').request_dt.strftime("%Y.%m.%d"))
                    print(data)
                    data.append(order_info.delivery_dt.strftime("%Y.%m.%d"))
                    print(data)
                    data.append(order_info.client.trainer.gym.name)
                    print(data)
                    data.append(order_info.client.trainer.name)
                    print(data)
                    data.append(order_info.client.name)
                    print(data)
                    try:
                        data.append(Delivery.objects.get(client_id=order_info.client.client_id).address)
                    except:
                        data.append('')
                    print(data)
                    try:
                        data.append(Delivery.objects.get(client_id=order_info.client.client_id).address_detail)
                    except:
                        data.append('')
                    print(data)

                    try:
                        if Delivery.objects.get(client_id=order_info.client.client_id).doorlock == 0:
                            doorlock = ''
                        else:
                            doorlock = Delivery.objects.get(client_id=order_info.client.client_id).doorlock

                        data.append(doorlock)
                    except:
                        data.append('')
                    print(data)
                    default_contact = str(order_info.client.contact)

                    if default_contact[0] == '1':
                        contact = '0' + default_contact[:2] + '-' + default_contact[2:7] + '-' + default_contact[7:]
                    else:
                        if len(default_contact) == 10:
                            contact = default_contact[:3] + '-' + default_contact[3:7] + '-' + default_contact[7:]
                        elif len(default_contact) < 10:
                            contact = default_contact[:3] + '-' + default_contact[3:6] + '-' + default_contact[6:]
                        elif len(default_contact) == 8 and order_info.client.contact[0] == '2':
                            contact = '0' + default_contact[:1] + '-' + default_contact[1:4] + '-' + default_contact[4:]
                        else:
                            contact = default_contact[:3] + '-' + default_contact[3:6] + '-' + default_contact[6:]

                    data.append(contact)
                    print(data)
                    data.append(order_info.default_amount)
                    print(data)
                    data.append(order_info.added_amount)
                    print(data)
                    data.append(order_info.total_amount)
                    print(data)

                    if order_info.is_delivery == True:
                        pickup = '배송'
                    else:
                        pickup = '픽업'

                    data.append(pickup)
                    print(data)
                    data.append(int(week_cnt)+1)
                    print(data)
                    data.append(order_info.day_cnt)
                    print(data)
                    data.append(Day(day_info.day).label)
                    print(data)

                    date = order_info.delivery_dt + timedelta(days=day_cnt)

                    data.append(date.strftime("%Y.%m.%d"))
                    print(data)
                    if week_info.meal_group == None or week_info.meal_group == '':
                        meal_group = ''
                    else:
                        meal_group = week_info.meal_group
                    data.append(meal_group)
                    print(data)
                    data.append(day_info.base_util.code)
                    print(data)
                    data.append(day_info.pro_util.code)
                    print(data)
                    data.append(day_info.veg_util.code)
                    print(data)
                    data.append(day_info.flavor_util.code)
                    print(data)
                    
                    try:
                        add_pro_list = Add_Pro.objects.filter(order_detail=day_info)
                        for add_pro_info in add_pro_list:
                            data.append(Add_Pro.objects.get(add_pro_id=add_pro_info.add_pro_id).pro_util.code)
                        if len(add_pro_list) < 1:
                            data.append('')
                            data.append('')
                        elif len(add_pro_list) < 2:
                            data.append('')
                    except:
                        pass

                    print(data)
                    
                    try:
                        add_veg_list = Add_Veg.objects.filter(order_detail=day_info)
                        for add_veg_info in add_veg_list:
                            data.append(Add_Veg.objects.get(add_veg_id=add_veg_info.add_veg_id).veg_util.code)
                        if len(add_veg_list) < 1:
                            data.append('')
                            data.append('')
                        elif len(add_veg_list) < 2:
                            data.append('')
                    except:
                        pass

                    print(data)
                    
                    try:
                        data.append(Add_Flavor.objects.get(order_detail=day_info).flavor_util.code)
                    except:
                        pass

                    print(data)

                    worksheet.append(data)
        import os

        output = io.BytesIO()
        # workbook.save(output)
        workbook.save(output)
        output.seek(0)

        return output

    # def save_excel_form(self, **kwargs):
    #     """   
    #     Parameters
    #     ----------
    #     kwargs
    #         불러와진 페이지에서 전달 받은 데이터
        
    #     Note
    #     ----
    #     대량 업로드를 위한 엑셀 폼을 다운받는 함수
    #     """
    #     index_list = kwargs.get('index').split(',')[:-1]

    #     output = BytesIO()
    #     workbook = xlsxwriter.Workbook(output)
    #     worksheet = workbook.add_worksheet('고객 목록')
    #     bold = workbook.add_format({'bold': True})
    #     columns = index_list

    #     row = 0
    #     for idx, val in enumerate(columns):
    #         worksheet.write(row, idx, val, bold)

    #     workbook.close()
    #     output.seek(0)
        
    #     return output

    # def save_bulk_data(self, **kwargs):
    #     """   
    #     Parameters
    #     ----------
    #     kwargs
    #         불러와진 페이지에서 전달 받은 데이터
    
    #     Note
    #     ----
    #     넘어온 데이터로 대량 업로드를 하는 함수
    #     """
    #     Logger.print_main_log('Starting Bulk Uploading client in Client page')
        
    #     ext = kwargs.get('file').name.split('.')[-1]

    #     if 'csv' in ext:
    #         test = pd.read_csv(kwargs.get('file'), encoding='cp949')
    #     else:
    #         test = pd.read_excel(kwargs.get('file'))

    #     for list in test.values.tolist():
    #         client_id = uuid.uuid4()
    #         Client(id=client_id, client_rank=str(list[0]), short_manager=str(list[1]), company=str(list[2]), manager=str(list[3]), email=str(list[7])).save()


    #         if str(list[4]) != 'nan':
    #             try:
    #                 material_id = Material.objects.get(name=str(list[4])).id
    #                 Client.objects.filter(id=client_id).update(material_id=material_id)
    #             except:
    #                 material_id = uuid.uuid4()
    #                 Material(id=material_id, name=str(list[4])).save()
    #                 Client.objects.filter(id=client_id).update(material_id=material_id)


    #         if str(list[5]) != 'nan':
    #             try:
    #                 purity_id = Purity.objects.get(purity=str(list[5])).id
    #                 Client.objects.filter(id=client_id).update(purity_id=purity_id)
    #             except:
    #                 purity_id = uuid.uuid4()
    #                 Purity(id=purity_id, purity=str(list[5])).save()
    #                 Client.objects.filter(id=client_id).update(purity_id=purity_id)

    #         if str(list[6]) != 'nan':
    #             try:
    #                 particle_id = Particle.objects.get(particle_size=str(list[6])).id
    #                 Client.objects.filter(id=client_id).update(particle_id=particle_id)

    #             except:
    #                 particle_id = uuid.uuid4()
    #                 Particle(id=particle_id, particle_size=str(list[6])).save()
    #                 Client.objects.filter(id=client_id).update(particle_id=particle_id)

    #     Logger.print_log('Successfully Bulk Uploaded client in Client page')

